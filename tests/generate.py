from openpyxl import Workbook
import random

def generate_excel(m,n,maxfit,maxpref,filename="output.xlsx"):
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Fit"
    sheet2 = wb.create_sheet(title="Preference")

    def fill_sheet(sheet,maxval):
        for col in range(2, n+2):
            sheet.cell(row=1,column=col, value=f"Project {col-1}")

        for row in range(2,m+2):
            sheet.cell(row=row, column=1, value=f"Team {row-1}")

        for row in range(2,m+2):
            for col in range(2,n+2):
                sheet.cell(row=row, column=col, value=random.randint(0,maxval))

    fill_sheet(sheet1,maxfit)
    fill_sheet(sheet2,maxpref)

    wb.save(filename)

random.seed("T039")
generate_excel(5,5,8,7,"Sample1.xlsx")
generate_excel(20,20,8,7,"Sample2.xlsx")
generate_excel(20,20,8,7,"Sample3.xlsx")
generate_excel(50,50,8,7,"Sample4.xlsx")
generate_excel(50,50,8,7,"Sample5.xlsx")
generate_excel(50,50,8,7,"Sample6.xlsx")
generate_excel(100,100,8,7,"Sample7.xlsx")
generate_excel(100,100,8,7,"Sample8.xlsx")
generate_excel(100,100,8,7,"Sample9.xlsx")
generate_excel(100,100,8,7,"Sample10.xlsx")
generate_excel(100,100,10,10,"Sample11.xlsx")
generate_excel(100,100,10,10,"Sample12.xlsx")
generate_excel(100,100,15,40,"Sample13.xlsx")
generate_excel(100,100,5,100,"Sample14.xlsx")
generate_excel(100,100,100,100,"Sample15.xlsx")